"""Phase 4 AppointmentService tests.

Integration-style: builds real Excel-backed repositories against a fresh
TEMPORARY COPY of the workbook for every test (never the original
template), per the Phase 4 test strategy. Verifies business rules and,
where a mutation occurs, that the underlying repository/workbook state
actually changed.

Some scenarios (an inactive doctor) cannot be constructed through the
documented repository interfaces, since DoctorRepository has no write
method — those tests manipulate the temporary copy directly with
openpyxl as test *setup* only, never touching the original template.
"""

import shutil
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import pytest

from app.api.schemas.appointment import (
    AppointmentApproval,
    AppointmentCancel,
    AppointmentCreate,
    AppointmentRejection,
    AppointmentStatus,
    AppointmentUpdate,
)
from app.api.schemas.availability import AvailabilityRequest
from app.repositories.appointment_repository import ExcelAppointmentRepository
from app.repositories.availability_repository import ExcelAvailabilityRepository
from app.repositories.doctor_repository import ExcelDoctorRepository
from app.services.appointment_service import AppointmentService, StaffContext
from app.services.exceptions import (
    AppointmentNotFound,
    DoctorInactive,
    DoctorNotFound,
    InvalidAppointmentState,
    InvalidSlotTime,
    OutsideAvailability,
    SlotUnavailable,
    UnauthorizedAction,
)

ORIGINAL_TEMPLATE = (
    Path(__file__).resolve().parents[1] / "data" / "clinic_appointments_MVP_template.xlsx"
)

# 2026-08-16 is a verified Sunday, matching the real workbook's demo data.
SUNDAY = date(2026, 8, 16)
MONDAY = date(2026, 8, 17)


@pytest.fixture()
def excel_copy(tmp_path) -> Path:
    dest = tmp_path / "test_workbook.xlsx"
    shutil.copy(ORIGINAL_TEMPLATE, dest)
    return dest


def _make_service(excel_copy, require_staff_approval=True, max_alternative_slots=3):
    return AppointmentService(
        appointment_repo=ExcelAppointmentRepository(excel_path=excel_copy),
        doctor_repo=ExcelDoctorRepository(excel_path=excel_copy),
        availability_repo=ExcelAvailabilityRepository(excel_path=excel_copy),
        require_staff_approval=require_staff_approval,
        max_alternative_slots=max_alternative_slots,
    )


def _deactivate_doctor(excel_copy: Path, doctor_id: str) -> None:
    """Test-setup only: DoctorRepository has no documented write method,
    so an inactive-doctor scenario is arranged by editing the temporary
    copy directly."""
    wb = openpyxl.load_workbook(excel_copy)
    ws = wb["Doctors"]
    header = [c.value for c in ws[1]]
    id_col = header.index("doctor_id")
    active_col = header.index("active")
    for row in ws.iter_rows(min_row=2):
        if row[id_col].value == doctor_id:
            row[active_col].value = False
    wb.save(excel_copy)


STAFF = StaffContext(is_staff=True, staff_id="staff-001")
NOT_STAFF = StaffContext(is_staff=False)


# --- check_availability -------------------------------------------------------

def test_check_availability_success(excel_copy):
    service = _make_service(excel_copy)
    resp = service.check_availability(
        AvailabilityRequest(doctor_id="DOC-001", appointment_date=SUNDAY, appointment_time=time(16, 0))
    )
    assert resp.available is True
    assert resp.message == "The requested slot is available."


def test_check_availability_doctor_not_found(excel_copy):
    service = _make_service(excel_copy)
    resp = service.check_availability(
        AvailabilityRequest(doctor_id="DOC-999", appointment_date=SUNDAY, appointment_time=time(16, 0))
    )
    assert resp.available is False
    assert resp.message == "Doctor not found."


def test_check_availability_inactive_doctor(excel_copy):
    _deactivate_doctor(excel_copy, "DOC-002")
    service = _make_service(excel_copy)
    resp = service.check_availability(
        AvailabilityRequest(doctor_id="DOC-002", appointment_date=SUNDAY, appointment_time=time(17, 0))
    )
    assert resp.available is False
    assert resp.message == "Doctor is not currently active."


def test_check_availability_unavailable_day(excel_copy):
    service = _make_service(excel_copy)
    resp = service.check_availability(
        AvailabilityRequest(doctor_id="DOC-001", appointment_date=MONDAY, appointment_time=time(16, 0))
    )
    assert resp.available is False
    assert resp.message == "No availability configured for the requested day."


def test_check_availability_outside_hours(excel_copy):
    service = _make_service(excel_copy)
    resp = service.check_availability(
        AvailabilityRequest(doctor_id="DOC-001", appointment_date=SUNDAY, appointment_time=time(20, 30))
    )
    assert resp.available is False
    assert resp.message == "Requested time is outside the doctor's available hours."


def test_check_availability_invalid_slot_time(excel_copy):
    service = _make_service(excel_copy)
    resp = service.check_availability(
        AvailabilityRequest(doctor_id="DOC-001", appointment_date=SUNDAY, appointment_time=time(16, 15))
    )
    assert resp.available is False
    assert resp.message == "Requested time does not align with the doctor's slot schedule."


def test_check_availability_conflict(excel_copy):
    service = _make_service(excel_copy)
    # APT-001 already occupies DOC-001 / Sunday / 17:00 with status Pending.
    resp = service.check_availability(
        AvailabilityRequest(doctor_id="DOC-001", appointment_date=SUNDAY, appointment_time=time(17, 0))
    )
    assert resp.available is False
    assert resp.message == "The requested slot is already booked."


# --- create_appointment ------------------------------------------------------

def test_create_appointment_success_pending_when_approval_required(excel_copy):
    service = _make_service(excel_copy, require_staff_approval=True)
    created = service.create_appointment(
        AppointmentCreate(
            patient_name="New Patient",
            patient_phone="03001112222",
            doctor_id="DOC-002",
            doctor_name="Wrong Name On Purpose",
            service="Dermatology Consultation",
            appointment_date=SUNDAY,
            appointment_time=time(17, 30),
        )
    )
    assert created.status == AppointmentStatus.PENDING
    assert created.doctor_name == "Dr. Sara"  # resolved from repository, not client input

    fetched = service.get_appointment(created.appointment_id)
    assert fetched.patient_name == "New Patient"


def test_create_appointment_auto_confirms_when_policy_disabled(excel_copy):
    service = _make_service(excel_copy, require_staff_approval=False)
    created = service.create_appointment(
        AppointmentCreate(
            patient_name="Auto Confirm Patient",
            patient_phone="03003334444",
            doctor_id="DOC-002",
            doctor_name="Dr. Sara",
            service="Dermatology Consultation",
            appointment_date=SUNDAY,
            appointment_time=time(18, 30),
        )
    )
    assert created.status == AppointmentStatus.CONFIRMED


def test_create_appointment_doctor_not_found_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(DoctorNotFound):
        service.create_appointment(
            AppointmentCreate(
                patient_name="Ghost Patient",
                patient_phone="0000000000",
                doctor_id="DOC-999",
                doctor_name="Nobody",
                service="General Consultation",
                appointment_date=SUNDAY,
                appointment_time=time(16, 0),
            )
        )


def test_create_appointment_inactive_doctor_raises(excel_copy):
    _deactivate_doctor(excel_copy, "DOC-002")
    service = _make_service(excel_copy)
    with pytest.raises(DoctorInactive):
        service.create_appointment(
            AppointmentCreate(
                patient_name="Patient",
                patient_phone="0000000000",
                doctor_id="DOC-002",
                doctor_name="Dr. Sara",
                service="Dermatology Consultation",
                appointment_date=SUNDAY,
                appointment_time=time(17, 0),
            )
        )


def test_create_appointment_conflict_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(SlotUnavailable):
        service.create_appointment(
            AppointmentCreate(
                patient_name="Conflict Patient",
                patient_phone="0000000000",
                doctor_id="DOC-001",
                doctor_name="Dr. Ahmed",
                service="General Consultation",
                appointment_date=SUNDAY,
                appointment_time=time(17, 0),  # already occupied by APT-001
            )
        )


# --- get_appointment ----------------------------------------------------------

def test_get_appointment_existing(excel_copy):
    service = _make_service(excel_copy)
    appt = service.get_appointment("APT-001")
    assert appt.patient_name == "Demo Patient"


def test_get_appointment_missing_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(AppointmentNotFound):
        service.get_appointment("APT-DOES-NOT-EXIST")


# --- update_appointment --------------------------------------------------------

def test_update_notes_only_no_availability_recheck_needed(excel_copy):
    service = _make_service(excel_copy)
    updated = service.update_appointment("APT-001", AppointmentUpdate(notes="Updated via test"))
    assert updated.notes == "Updated via test"
    assert updated.appointment_time == time(17, 0)  # unchanged


def test_update_reschedule_success(excel_copy):
    service = _make_service(excel_copy)
    updated = service.update_appointment(
        "APT-002", AppointmentUpdate(appointment_time=time(16, 30))
    )
    assert updated.appointment_time == time(16, 30)
    assert updated.doctor_name == "Dr. Ahmed"


def test_update_reschedule_conflict_raises(excel_copy):
    service = _make_service(excel_copy)
    # Rescheduling APT-002 into APT-001's occupied slot.
    with pytest.raises(SlotUnavailable):
        service.update_appointment("APT-002", AppointmentUpdate(appointment_time=time(17, 0)))


def test_update_locked_state_raises(excel_copy):
    service = _make_service(excel_copy)
    service.cancel_appointment("APT-001", AppointmentCancel())
    with pytest.raises(InvalidAppointmentState):
        service.update_appointment("APT-001", AppointmentUpdate(notes="should fail"))


def test_update_missing_appointment_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(AppointmentNotFound):
        service.update_appointment("APT-DOES-NOT-EXIST", AppointmentUpdate(notes="x"))


def test_update_outside_hours_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(OutsideAvailability):
        service.update_appointment("APT-001", AppointmentUpdate(appointment_time=time(21, 0)))


# --- cancel_appointment ---------------------------------------------------------

def test_cancel_success(excel_copy):
    service = _make_service(excel_copy)
    cancelled = service.cancel_appointment("APT-001", AppointmentCancel(reason="Patient request"))
    assert cancelled.status == AppointmentStatus.CANCELLED

    # Cancelled appointment must not block its former slot.
    resp = service.check_availability(
        AvailabilityRequest(doctor_id="DOC-001", appointment_date=SUNDAY, appointment_time=time(17, 0))
    )
    assert resp.available is True


def test_cancel_already_cancelled_raises(excel_copy):
    service = _make_service(excel_copy)
    service.cancel_appointment("APT-001", AppointmentCancel())
    with pytest.raises(InvalidAppointmentState):
        service.cancel_appointment("APT-001", AppointmentCancel())


def test_cancel_missing_appointment_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(AppointmentNotFound):
        service.cancel_appointment("APT-DOES-NOT-EXIST", AppointmentCancel())


# --- approve_appointment ----------------------------------------------------------

def test_approve_success(excel_copy):
    service = _make_service(excel_copy)
    approved = service.approve_appointment("APT-001", AppointmentApproval(), STAFF)
    assert approved.status == AppointmentStatus.CONFIRMED


def test_approve_unauthorized_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(UnauthorizedAction):
        service.approve_appointment("APT-001", AppointmentApproval(), NOT_STAFF)


def test_approve_invalid_state_raises(excel_copy):
    service = _make_service(excel_copy)
    # APT-002 is already Confirmed in the template.
    with pytest.raises(InvalidAppointmentState):
        service.approve_appointment("APT-002", AppointmentApproval(), STAFF)


def test_approve_missing_appointment_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(AppointmentNotFound):
        service.approve_appointment("APT-DOES-NOT-EXIST", AppointmentApproval(), STAFF)


# --- reject_appointment -----------------------------------------------------------

def test_reject_success(excel_copy):
    service = _make_service(excel_copy)
    rejected = service.reject_appointment(
        "APT-001", AppointmentRejection(reason="Doctor unavailable"), STAFF
    )
    assert rejected.status == AppointmentStatus.REJECTED


def test_reject_invalid_state_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(InvalidAppointmentState):
        service.reject_appointment("APT-002", AppointmentRejection(reason="Already confirmed"), STAFF)


def test_reject_unauthorized_raises(excel_copy):
    service = _make_service(excel_copy)
    with pytest.raises(UnauthorizedAction):
        service.reject_appointment("APT-001", AppointmentRejection(reason="No permission"), NOT_STAFF)


def test_reject_reason_too_short_rejected_at_schema_level():
    # Enforced by the Phase 2 Pydantic schema (min_length=3); the service
    # never even receives an invalid AppointmentRejection instance.
    from pydantic import ValidationError

    with pytest.raises(ValidationError):
        AppointmentRejection(reason="no")


# --- find_alternative_slots --------------------------------------------------------

def test_find_alternative_slots_when_unavailable(excel_copy):
    service = _make_service(excel_copy)
    resp = service.find_alternative_slots(
        AvailabilityRequest(doctor_id="DOC-001", appointment_date=SUNDAY, appointment_time=time(17, 0))
    )
    assert resp.requested_slot_available is False
    assert len(resp.alternatives) > 0
    occupied = {time(17, 0), time(18, 0)}
    for alt in resp.alternatives:
        assert alt.appointment_time not in occupied
        assert alt.doctor_id == "DOC-001"
        assert alt.doctor_name == "Dr. Ahmed"


def test_find_alternative_slots_respects_max_count(excel_copy):
    service = _make_service(excel_copy, max_alternative_slots=2)
    resp = service.find_alternative_slots(
        AvailabilityRequest(doctor_id="DOC-001", appointment_date=SUNDAY, appointment_time=time(17, 0))
    )
    assert len(resp.alternatives) <= 2


def test_find_alternative_slots_inactive_doctor_returns_no_fabricated_slots(excel_copy):
    _deactivate_doctor(excel_copy, "DOC-002")
    service = _make_service(excel_copy)
    resp = service.find_alternative_slots(
        AvailabilityRequest(doctor_id="DOC-002", appointment_date=SUNDAY, appointment_time=time(17, 0))
    )
    assert resp.requested_slot_available is False
    assert resp.alternatives == []
