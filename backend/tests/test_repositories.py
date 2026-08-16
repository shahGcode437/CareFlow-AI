"""Phase 3 repository tests.

All tests operate against a temporary COPY of
data/clinic_appointments_MVP_template.xlsx, made fresh for every test via
the `excel_copy` fixture below. The original template is never opened
for writing by any test in this file.

Scope: repository return values AND actual workbook persistence for
mutating operations, per the Phase 3 test strategy. Business rules
(state transitions, availability decisions, approval/rejection policy)
are explicitly out of scope for this phase and are not tested here.
"""

import shutil
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import pytest

from app.api.schemas.appointment import AppointmentResponse, AppointmentStatus
from app.repositories.appointment_repository import ExcelAppointmentRepository
from app.repositories.audit_repository import ExcelAuditRepository
from app.repositories.availability_repository import ExcelAvailabilityRepository
from app.repositories.doctor_repository import ExcelDoctorRepository

ORIGINAL_TEMPLATE = Path(__file__).resolve().parents[1] / "data" / "clinic_appointments_MVP_template.xlsx"


@pytest.fixture()
def excel_copy(tmp_path) -> Path:
    """Fresh temporary copy of the real workbook for one test only."""
    dest = tmp_path / "test_workbook.xlsx"
    shutil.copy(ORIGINAL_TEMPLATE, dest)
    return dest


@pytest.fixture()
def original_template_snapshot():
    """Record the original template's bytes before the test session-ish
    scope so we can assert it was never modified."""
    return ORIGINAL_TEMPLATE.read_bytes()


# --- Safety: original template must never be modified -----------------------

def test_original_template_untouched_by_a_mutating_workflow(excel_copy, original_template_snapshot):
    # Perform a real mutation against the COPY only.
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    new_appt = AppointmentResponse(
        appointment_id="APT-999",
        patient_name="Safety Check Patient",
        patient_phone="03001234567",
        doctor_id="DOC-001",
        doctor_name="Dr. Ahmed",
        service="General Consultation",
        appointment_date=date(2026, 8, 23),
        appointment_time=time(16, 0),
        status=AppointmentStatus.PENDING,
        created_at=datetime(2026, 8, 20, 9, 0),
        updated_at=datetime(2026, 8, 20, 9, 0),
        notes=None,
    )
    repo.create(new_appt)

    assert ORIGINAL_TEMPLATE.read_bytes() == original_template_snapshot, (
        "The original template must never be modified by tests."
    )


# --- AppointmentRepository ----------------------------------------------------

def test_get_by_id_existing(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    appt = repo.get_by_id("APT-001")
    assert appt is not None
    assert appt.patient_name == "Demo Patient"
    assert appt.doctor_id == "DOC-001"
    assert appt.status == AppointmentStatus.PENDING
    assert appt.appointment_date == date(2026, 8, 16)
    assert appt.appointment_time == time(17, 0)


def test_get_by_id_missing_returns_none(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    assert repo.get_by_id("APT-DOES-NOT-EXIST") is None


def test_list_by_doctor_and_date(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    results = repo.list_by_doctor_and_date("DOC-001", date(2026, 8, 16))
    ids = {a.appointment_id for a in results}
    assert ids == {"APT-001", "APT-002"}


def test_list_by_doctor_and_date_no_match(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    results = repo.list_by_doctor_and_date("DOC-001", date(2099, 1, 1))
    assert results == []


def test_create_appointment_persists_to_workbook(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    new_appt = AppointmentResponse(
        appointment_id="APT-100",
        patient_name="New Patient",
        patient_phone="03009999999",
        doctor_id="DOC-002",
        doctor_name="Dr. Sara",
        service="Dermatology Consultation",
        appointment_date=date(2026, 8, 23),
        appointment_time=time(17, 30),
        status=AppointmentStatus.PENDING,
        created_at=datetime(2026, 8, 20, 9, 0),
        updated_at=datetime(2026, 8, 20, 9, 0),
        notes="Created by repository test",
    )
    returned = repo.create(new_appt)
    assert returned.appointment_id == "APT-100"

    # Repository-level read confirms it.
    fetched = repo.get_by_id("APT-100")
    assert fetched is not None
    assert fetched.patient_name == "New Patient"

    # Direct workbook inspection confirms actual persistence, independent
    # of the repository's own read path.
    wb = openpyxl.load_workbook(excel_copy)
    ws = wb["Appointments"]
    header = [c.value for c in ws[1]]
    id_col = header.index("appointment_id")
    ids_in_sheet = [row[id_col].value for row in ws.iter_rows(min_row=2)]
    assert "APT-100" in ids_in_sheet


def test_update_appointment_persists_to_workbook(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    existing = repo.get_by_id("APT-001")
    assert existing is not None

    updated = existing.model_copy(
        update={
            "appointment_time": time(19, 0),
            "status": AppointmentStatus.CONFIRMED,
            "updated_at": datetime(2026, 8, 20, 10, 0),
        }
    )
    result = repo.update(updated)
    assert result.status == AppointmentStatus.CONFIRMED

    refetched = repo.get_by_id("APT-001")
    assert refetched.appointment_time == time(19, 0)
    assert refetched.status == AppointmentStatus.CONFIRMED

    # Original demo row for APT-002 must be untouched.
    other = repo.get_by_id("APT-002")
    assert other.status == AppointmentStatus.CONFIRMED  # was already Confirmed in template
    assert other.appointment_time == time(18, 0)


def test_update_nonexistent_appointment_raises(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    ghost = AppointmentResponse(
        appointment_id="APT-GHOST",
        patient_name="Ghost",
        patient_phone="0000000000",
        doctor_id="DOC-001",
        doctor_name="Dr. Ahmed",
        service="General Consultation",
        appointment_date=date(2026, 8, 16),
        appointment_time=time(17, 0),
        status=AppointmentStatus.CONFIRMED,
        created_at=datetime(2026, 8, 20, 9, 0),
        updated_at=datetime(2026, 8, 20, 9, 0),
    )
    with pytest.raises(ValueError):
        repo.update(ghost)


def test_exists_conflict_true_for_active_appointment(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    # APT-001 occupies DOC-001 / 2026-08-16 / 17:00, status Pending (blocking).
    assert repo.exists_conflict("DOC-001", date(2026, 8, 16), time(17, 0)) is True


def test_exists_conflict_false_for_open_slot(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    assert repo.exists_conflict("DOC-001", date(2026, 8, 16), time(16, 0)) is False


def test_exists_conflict_excludes_given_appointment_id(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    # Excluding APT-001 itself from the conflict check for its own slot.
    assert (
        repo.exists_conflict(
            "DOC-001", date(2026, 8, 16), time(17, 0), exclude_appointment_id="APT-001"
        )
        is False
    )


def test_exists_conflict_cancelled_does_not_block(excel_copy):
    repo = ExcelAppointmentRepository(excel_path=excel_copy)
    existing = repo.get_by_id("APT-001")
    cancelled = existing.model_copy(update={"status": AppointmentStatus.CANCELLED})
    repo.update(cancelled)

    # Per Service Design §8: a cancelled appointment does not block the slot.
    assert repo.exists_conflict("DOC-001", date(2026, 8, 16), time(17, 0)) is False


# --- DoctorRepository -----------------------------------------------------------

def test_doctor_get_by_id(excel_copy):
    repo = ExcelDoctorRepository(excel_path=excel_copy)
    doc = repo.get_by_id("DOC-001")
    assert doc is not None
    assert doc.doctor_name == "Dr. Ahmed"
    assert doc.specialty == "General Medicine"
    assert doc.active is True


def test_doctor_get_by_id_missing(excel_copy):
    repo = ExcelDoctorRepository(excel_path=excel_copy)
    assert repo.get_by_id("DOC-DOES-NOT-EXIST") is None


def test_doctor_list_active(excel_copy):
    repo = ExcelDoctorRepository(excel_path=excel_copy)
    doctors = repo.list_active()
    ids = {d.doctor_id for d in doctors}
    assert ids == {"DOC-001", "DOC-002"}  # both demo doctors are active


# --- AvailabilityRepository -----------------------------------------------------

def test_availability_get_for_doctor(excel_copy):
    repo = ExcelAvailabilityRepository(excel_path=excel_copy)
    rules = repo.get_for_doctor("DOC-001")
    assert len(rules) == 1
    assert rules[0].day_of_week == "Sunday"
    assert rules[0].start_time == time(16, 0)
    assert rules[0].end_time == time(20, 0)
    assert rules[0].slot_duration_minutes == 30


def test_availability_get_for_day(excel_copy):
    repo = ExcelAvailabilityRepository(excel_path=excel_copy)
    rules = repo.get_for_day("DOC-002", "Sunday")
    assert len(rules) == 1
    assert rules[0].availability_id == "AVL-002"

    no_match = repo.get_for_day("DOC-002", "Monday")
    assert no_match == []


# --- AuditRepository (flagged / partial) -----------------------------------------

def test_audit_create_persists_known_columns(excel_copy):
    repo = ExcelAuditRepository(excel_path=excel_copy)
    repo.create(
        {
            "event_id": "EVT-TEST-001",
            "appointment_id": "APT-001",
            "action": "CREATE",
            "actor_type": "test",
            "actor_id": "pytest",
            "timestamp": datetime(2026, 8, 20, 9, 0),
            "old_status": None,
            "new_status": "Pending",
            "reason": "Repository test",
        }
    )

    # Direct workbook inspection (not the repository) confirms persistence,
    # since AuditRepository documents no read method.
    wb = openpyxl.load_workbook(excel_copy)
    ws = wb["Audit_Log"]
    header = [c.value for c in ws[1]]
    event_id_col = header.index("event_id")
    event_ids = [row[event_id_col].value for row in ws.iter_rows(min_row=2)]
    assert "EVT-TEST-001" in event_ids


def test_audit_create_rejects_unknown_column(excel_copy):
    repo = ExcelAuditRepository(excel_path=excel_copy)
    with pytest.raises(ValueError):
        repo.create({"summary": "not a real column", "metadata": "also not real"})
