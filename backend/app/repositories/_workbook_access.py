"""Internal helper: generic, sheet-agnostic Excel row access.

Provides the low-level open/read/append/update/save mechanics shared by
all four Excel-backed repositories, plus the single application-level
write lock required by Service Design §19 ("Excel Concurrency / Safe
Write Strategy") and Implementation Guide Phase 3 ("Use an
application-level write lock for Excel mutation workflows").

This module is a private implementation detail of the repository layer
(underscore-prefixed, not part of the documented repository contract).
It knows nothing about appointments, doctors, or business rules — only
"a worksheet has a header row and data rows, keyed by column name."
"""

import threading
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook

# A single process-wide lock protecting the read-modify-write-save
# sequence for ANY mutation to the workbook. This is the "single write
# boundary" called for by Service Design §19-20. Appropriate for MVP/
# local single-process development; not a substitute for a real
# database transaction (documented explicitly in Service Design §20).
WRITE_LOCK = threading.Lock()


def _load_workbook(path: Path) -> Workbook:
    if not path.exists():
        raise FileNotFoundError(
            f"Excel workbook not found at {path}. Check the EXCEL_FILE_PATH "
            "setting (see app/core/config.py)."
        )
    return openpyxl.load_workbook(path)


def read_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read all data rows (below the header) from a sheet as a list of
    dicts keyed by the sheet's actual header row values.

    Read-only: does not require the write lock.
    """
    wb = _load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue  # skip fully blank trailing rows
        rows.append(dict(zip(header, row)))
    return rows


def append_row(path: Path, sheet_name: str, row_values: dict[str, Any]) -> None:
    """Append one row to a sheet, writing values in the sheet's own
    header-column order. Caller must hold WRITE_LOCK.
    """
    wb = _load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    ordered = [row_values.get(col) for col in header]
    ws.append(ordered)
    wb.save(path)


def update_row_by_key(
    path: Path,
    sheet_name: str,
    key_column: str,
    key_value: Any,
    row_values: dict[str, Any],
) -> bool:
    """Find the row whose `key_column` equals `key_value` and overwrite
    only the columns present in `row_values`, leaving other columns
    untouched. Returns True if a row was found and updated, False
    otherwise. Caller must hold WRITE_LOCK.
    """
    wb = _load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    if key_column not in header:
        raise ValueError(f"Key column '{key_column}' not found in sheet '{sheet_name}'.")
    key_idx = header.index(key_column)

    for row in ws.iter_rows(min_row=2):
        if row[key_idx].value == key_value:
            for col_name, value in row_values.items():
                if col_name not in header:
                    continue
                col_idx = header.index(col_name)
                row[col_idx].value = value
            wb.save(path)
            return True
    return False
